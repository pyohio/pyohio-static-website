"""Sponsor data commands."""

from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import click
import httpx
from openpyxl import load_workbook
from ruamel.yaml import YAML

DEFAULT_PAGE = Path("2026/content/sponsors/individual-sponsors.md")
DEFAULT_MIN_AMOUNT = 50.0

SHEET_ID_RE = re.compile(r"/d/([A-Za-z0-9_-]+)")


@dataclass
class DonorAggregate:
    name: str
    total: float
    first_date: datetime


@click.group()
def sponsors():
    """Sponsor data commands."""


@sponsors.command("update-individual")
@click.option(
    "--sheet-url",
    envvar="INDIVIDUAL_SPONSORS_SHEET_URL",
    help="Google Sheets share/edit/export URL (or set INDIVIDUAL_SPONSORS_SHEET_URL).",
)
@click.option(
    "--xlsx-file",
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
    help="Read from a local .xlsx instead of fetching from Google Sheets (for testing).",
)
@click.option(
    "--page",
    type=click.Path(dir_okay=False, path_type=Path),
    default=DEFAULT_PAGE,
    show_default=True,
    help="Markdown file whose frontmatter `sponsors:` key will be rewritten.",
)
@click.option(
    "--min-amount",
    type=float,
    default=DEFAULT_MIN_AMOUNT,
    show_default=True,
    help="Minimum aggregate donation per donor to be listed.",
)
@click.option(
    "--dry-run",
    is_flag=True,
    help="Print sorted names to stdout without writing.",
)
def update_individual(
    sheet_url: str | None,
    xlsx_file: Path | None,
    page: Path,
    min_amount: float,
    dry_run: bool,
):
    """Refresh the individual sponsors list from the PSF contributions sheet."""
    if not xlsx_file and not sheet_url:
        raise click.UsageError(
            "Provide --sheet-url (or INDIVIDUAL_SPONSORS_SHEET_URL) or --xlsx-file."
        )

    if xlsx_file:
        click.echo(f"Reading xlsx from {xlsx_file}", err=True)
        workbook_bytes = xlsx_file.read_bytes()
    else:
        export_url = _build_export_url(sheet_url)
        click.echo(f"Fetching sheet from {export_url}", err=True)
        response = httpx.get(export_url, follow_redirects=True, timeout=30.0)
        response.raise_for_status()
        workbook_bytes = response.content

    rows = _read_contribution_rows(workbook_bytes)
    click.echo(f"Loaded {len(rows)} contribution rows", err=True)

    donors = _aggregate_by_contact(rows)
    click.echo(f"Aggregated to {len(donors)} unique donors", err=True)

    eligible = [d for d in donors if d.total >= min_amount]
    eligible.sort(key=lambda d: (-d.total, d.first_date))
    click.echo(
        f"{len(eligible)} donors meet the ${min_amount:.0f} threshold", err=True
    )

    names = [d.name for d in eligible]

    if dry_run:
        for d in eligible:
            click.echo(f"  ${d.total:>8.2f}  {d.first_date.date()}  {d.name}")
        return

    if not page.exists():
        raise click.UsageError(f"Page file not found: {page}")

    _write_sponsors_frontmatter(page, names)
    click.echo(f"Wrote {len(names)} sponsors to {page}", err=True)


def _build_export_url(sheet_url: str) -> str:
    """Convert a share/edit/export URL into an xlsx export URL."""
    sheet_url = sheet_url.strip()
    match = SHEET_ID_RE.search(sheet_url)
    if not match:
        raise click.UsageError(f"Could not extract sheet ID from URL: {sheet_url!r}")
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def _read_contribution_rows(workbook_bytes: bytes) -> list[dict]:
    """Parse the first worksheet into row dicts.

    Expected columns (based on PSF CiviCRM contribution export):
      A Display Name, B First Name, C Last Name, D Contact ID,
      E Email, F Contribution Date, G Total Amount
    """
    workbook = load_workbook(
        io.BytesIO(workbook_bytes), read_only=True, data_only=True
    )
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    parsed: list[dict] = []
    for raw in rows[1:]:
        if not raw or len(raw) < 7:
            continue
        display_name, _first, _last, contact_id, _email, contrib_date, amount = raw[:7]
        if not display_name or contact_id is None or amount is None:
            continue
        if not isinstance(contrib_date, datetime):
            continue
        try:
            amount_f = float(amount)
        except (TypeError, ValueError):
            continue
        parsed.append(
            {
                "name": str(display_name).strip(),
                "contact_id": int(contact_id),
                "date": contrib_date,
                "amount": amount_f,
            }
        )
    return parsed


def _aggregate_by_contact(rows: list[dict]) -> list[DonorAggregate]:
    """Group rows by Contact ID, summing amount and tracking first donation date."""
    by_id: dict[int, dict] = defaultdict(
        lambda: {"name": "", "total": 0.0, "first_date": None}
    )
    for row in rows:
        bucket = by_id[row["contact_id"]]
        bucket["name"] = row["name"]
        bucket["total"] += row["amount"]
        if bucket["first_date"] is None or row["date"] < bucket["first_date"]:
            bucket["first_date"] = row["date"]

    return [
        DonorAggregate(name=v["name"], total=v["total"], first_date=v["first_date"])
        for v in by_id.values()
    ]


def _write_sponsors_frontmatter(page: Path, names: list[str]) -> None:
    """Rewrite the `sponsors:` key in the page's YAML frontmatter, preserving body."""
    text = page.read_text()
    if not text.startswith("---\n"):
        raise click.UsageError(f"{page} does not start with YAML frontmatter")
    end = text.find("\n---\n", 4)
    if end == -1:
        raise click.UsageError(f"{page} has no closing frontmatter delimiter")

    frontmatter_text = text[4:end]
    body = text[end + len("\n---\n") :]

    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.indent(mapping=2, sequence=4, offset=2)
    data = yaml.load(frontmatter_text) or {}
    data["sponsors"] = names

    buf = io.StringIO()
    yaml.dump(data, buf)
    new_frontmatter = buf.getvalue()
    if not new_frontmatter.endswith("\n"):
        new_frontmatter += "\n"

    page.write_text(f"---\n{new_frontmatter}---\n{body}")
